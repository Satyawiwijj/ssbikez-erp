import json
from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.utils import timezone

from accounts.audit import log_action
from accounts.permissions import require_module_action
from masters.models import Warehouse, Rack, Bin, Supplier
from django.views.decorators.http import require_POST

from .models import (
    SparesItem, ItemRackBin, StockLedger,
    SupplierQuote, SupplierQuoteItem,
    PurchaseOrder, PurchaseOrderItem,
    PurchaseInvoice, PurchaseInvoiceItem,
    CounterSale, CounterSaleItem,
    CounterSaleReturn, CounterSaleReturnItem,
    SparesIssueAlteration, SparesIssueAlterationItem, SparesIssueAlterationDeletedItem,
    StockTransfer, StockTransferItem,
    StockCountUpdate, StockCountItem,
    RequestSupplierQuote, RequestSupplierQuoteItem,
    SparesPurchaseEstimationMaster, SparesPurchaseEstimationItem, SparesPurchaseEstimationLabor,
    ServiceSparesIssueReturn, ServiceSparesIssueReturnItem,
    VehicleSparesMaster, SparesMRPPriceRevision, SparesMRPPriceRevisionItem,
    SparesProfitPercentageSettings, SparesPurchaseQtyDaysSettings,
    ServiceSparesWarranty, ServiceSparesWarrantyItem,
)
from .forms import (
    SparesItemForm,
    SupplierQuoteForm, SupplierQuoteItemFormSet, SupplierQuoteTaxFormSet,
    PurchaseOrderForm, PurchaseOrderItemFormSet, PurchaseOrderTaxFormSet,
    PurchaseInvoiceForm, PurchaseInvoiceItemFormSet, PurchaseInvoiceTaxFormSet,
    CounterSaleForm, CounterSaleItemFormSet,
    CounterSaleReturnForm, CounterSaleReturnItemFormSet,
    SparesIssueAlterationForm, SparesIssueAlterationItemFormSet,
    SparesIssueAlterationDeletedItemFormSet,
    StockTransferForm, StockTransferItemFormSet,
    StockCountUpdateForm, StockCountItemFormSet,
    RequestSupplierQuoteForm, RequestSupplierQuoteItemFormSet,
    SparesPurchaseEstimationMasterForm, SparesPurchaseEstimationItemFormSet,
    SparesPurchaseEstimationLaborFormSet,
    ServiceSparesIssueReturnForm, ServiceSparesIssueReturnItemFormSet,
    VehicleSparesMasterForm, SparesMRPPriceRevisionForm, SparesMRPPriceRevisionItemFormSet,
    SparesProfitPercentageSettingsForm, SparesPurchaseQtyDaysSettingsForm,
    ServiceSparesWarrantyForm, ServiceSparesWarrantyItemFormSet,
)


# -- Dashboard ----------------------------------------------------------------

@login_required
def dashboard(request):
    def safe_count(qs):
        try:
            return qs.count()
        except Exception:
            return 0

    total_items = safe_count(SparesItem.objects.filter(is_active=True))

    try:
        low_stock = SparesItem.objects.filter(
            maintain_stock=True, is_active=True, reorder_level__gt=0
        ).count()
    except Exception:
        low_stock = 0

    today = timezone.localdate()
    counter_sales_today = safe_count(CounterSale.objects.filter(date=today))
    pending_pos = safe_count(PurchaseOrder.objects.filter(status__in=['draft', 'submitted']))
    recent_sales = CounterSale.objects.select_related('godown').order_by('-created_at')[:8]

    return render(request, 'spares/dashboard.html', {
        'total_items': total_items,
        'low_stock': low_stock,
        'counter_sales_today': counter_sales_today,
        'pending_pos': pending_pos,
        'recent_sales': recent_sales,
    })


# -- Items --------------------------------------------------------------------

@login_required
def item_list(request):
    q = request.GET.get('q', '')
    items = SparesItem.objects.select_related('category')
    if q:
        items = items.filter(
            Q(item_name__icontains=q) | Q(item_code__icontains=q) |
            Q(part_number__icontains=q) | Q(hsn_sac__icontains=q)
        )
    paginator = Paginator(items, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/item_list.html', {'items': page_obj, 'page_obj': page_obj, 'q': q})


@login_required
def item_detail(request, pk):
    item = get_object_or_404(SparesItem, pk=pk)
    stock = item.stock.select_related('warehouse', 'rack', 'bin').all()
    return render(request, 'spares/item_detail.html', {'item': item, 'stock': stock})


@login_required
@require_module_action('spares', 'create')
def item_create(request):
    form = SparesItemForm(request.POST or None)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.created_by = request.user
        obj.save()
        messages.success(request, f'Item {obj.item_code} created.')
        return redirect('spares:item_detail', pk=obj.pk)
    return render(request, 'spares/item_form.html', {'form': form, 'title': 'New Item'})


@login_required
@require_module_action('spares', 'edit')
def item_update(request, pk):
    obj = get_object_or_404(SparesItem, pk=pk)
    form = SparesItemForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Item updated.')
        return redirect('spares:item_detail', pk=pk)
    return render(request, 'spares/item_form.html', {'form': form, 'title': 'Edit Item', 'object': obj})


# -- Stock Report -------------------------------------------------------------

@login_required
def stock_report(request):
    ledger = StockLedger.objects.select_related(
        'item', 'warehouse', 'rack', 'bin'
    ).filter(quantity__gt=0).order_by('warehouse__name', 'item__item_name')
    warehouses = Warehouse.objects.filter(is_active=True)
    selected_wh = request.GET.get('warehouse', '')
    if selected_wh:
        ledger = ledger.filter(warehouse_id=selected_wh)
    return render(request, 'spares/stock_report.html', {
        'ledger': ledger,
        'warehouses': warehouses,
        'selected_wh': selected_wh,
    })


# -- Supplier Quotes ----------------------------------------------------------

@login_required
def quote_list(request):
    quotes = SupplierQuote.objects.select_related('supplier').order_by('-date')
    paginator = Paginator(quotes, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/quote_list.html', {'quotes': page_obj, 'page_obj': page_obj})


@login_required
def quote_detail(request, pk):
    quote = get_object_or_404(SupplierQuote, pk=pk)
    items = quote.items.select_related('item')
    return render(request, 'spares/quote_detail.html', {'quote': quote, 'items': items})


@login_required
@require_module_action('spares', 'create')
def quote_create(request):
    initial = {}
    if request.GET.get('request_quotation'):
        initial['request_quotation'] = request.GET['request_quotation']
    form = SupplierQuoteForm(request.POST or None, initial=initial)
    formset = SupplierQuoteItemFormSet(request.POST or None, prefix='items')
    tax_formset = SupplierQuoteTaxFormSet(request.POST or None, prefix='taxes')
    if request.method == 'POST' and form.is_valid() and formset.is_valid() and tax_formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            formset.instance = obj
            formset.save()
            tax_formset.instance = obj
            tax_formset.save()
            items = obj.items.all()
            obj.total_quantity = sum(i.quantity for i in items)
            obj.total_amount = sum(i.amount for i in items)
            obj.total_taxes = sum(t.amount for t in obj.taxes.all())
            obj.grand_total = obj.total_amount - obj.additional_discount_amount + obj.total_taxes
            obj.save()
        messages.success(request, f'Quote {obj.quote_no} created.')
        return redirect('spares:quote_detail', pk=obj.pk)
    return render(request, 'spares/quote_form.html', {
        'form': form, 'formset': formset, 'tax_formset': tax_formset, 'title': 'New Supplier Quote'
    })


@login_required
@require_module_action('spares', 'edit')
def quote_update(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(SupplierQuote, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    form = SupplierQuoteForm(request.POST or None, instance=obj)
    formset = SupplierQuoteItemFormSet(request.POST or None, instance=obj, prefix='items')
    tax_formset = SupplierQuoteTaxFormSet(request.POST or None, instance=obj, prefix='taxes')
    if request.method == 'POST' and form.is_valid() and formset.is_valid() and tax_formset.is_valid():
        with transaction.atomic():
            form.save()
            formset.save()
            tax_formset.save()
            items = obj.items.all()
            obj.total_quantity = sum(i.quantity for i in items)
            obj.total_amount = sum(i.amount for i in items)
            obj.total_taxes = sum(t.amount for t in obj.taxes.all())
            obj.grand_total = obj.total_amount - obj.additional_discount_amount + obj.total_taxes
            obj.save()
        messages.success(request, 'Quote updated.')
        return redirect('spares:quote_detail', pk=pk)
    return render(request, 'spares/quote_form.html', {
        'form': form, 'formset': formset, 'tax_formset': tax_formset, 'title': 'Edit Quote', 'object': obj
    })


# -- Purchase Orders ----------------------------------------------------------

@login_required
def order_list(request):
    orders = PurchaseOrder.objects.select_related('supplier').order_by('-date')
    status_filter = request.GET.get('status', '')
    if status_filter:
        orders = orders.filter(status=status_filter)
    paginator = Paginator(orders, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/order_list.html', {
        'orders': page_obj,
        'page_obj': page_obj,
        'status_choices': PurchaseOrder.STATUS,
        'status_filter': status_filter,
    })


@login_required
def order_detail(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    items = order.items.select_related('item', 'warehouse')
    return render(request, 'spares/order_detail.html', {'order': order, 'items': items})


@login_required
@require_module_action('spares', 'create')
def order_create(request):
    initial = {}
    if request.GET.get('estimation'):
        initial['estimation'] = request.GET['estimation']
        initial['get_estimation'] = True
    if request.GET.get('customer_order'):
        initial['customer_order'] = request.GET['customer_order']
        initial['get_customer_order'] = True
    form = PurchaseOrderForm(request.POST or None, initial=initial)
    formset = PurchaseOrderItemFormSet(request.POST or None, prefix='items')
    tax_formset = PurchaseOrderTaxFormSet(request.POST or None, prefix='taxes')
    if request.method == 'POST' and form.is_valid() and formset.is_valid() and tax_formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            formset.instance = obj
            formset.save()
            tax_formset.instance = obj
            tax_formset.save()
            items = obj.items.all()
            obj.total_quantity = sum(i.quantity for i in items)
            obj.total_amount = sum(i.amount for i in items)
            obj.total_taxes = sum(t.amount for t in obj.taxes.all())
            obj.grand_total = obj.total_amount + obj.total_taxes
            obj.save()
        messages.success(request, f'PO {obj.po_no} created.')
        return redirect('spares:order_detail', pk=obj.pk)
    return render(request, 'spares/po_form.html', {
        'form': form, 'formset': formset, 'tax_formset': tax_formset, 'title': 'New Purchase Order'
    })


@login_required
@require_module_action('spares', 'edit')
def order_update(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(PurchaseOrder, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    # Edit-lock: once a PO is no longer Draft (Submitted/Received/Cancelled),
    # block the full formset rewrite this view otherwise allows -- previously
    # a Submitted PO stayed fully editable even after a Purchase Invoice had
    # already been created against it, letting the two silently diverge.
    if obj.status != 'draft':
        messages.error(
            request,
            f'{obj.po_no} is "{obj.get_status_display()}" and can no longer be edited. '
            'Only a Draft purchase order can be changed.'
        )
        return redirect('spares:order_detail', pk=pk)
    form = PurchaseOrderForm(request.POST or None, instance=obj)
    formset = PurchaseOrderItemFormSet(request.POST or None, instance=obj, prefix='items')
    tax_formset = PurchaseOrderTaxFormSet(request.POST or None, instance=obj, prefix='taxes')
    if request.method == 'POST' and form.is_valid() and formset.is_valid() and tax_formset.is_valid():
        with transaction.atomic():
            form.save()
            formset.save()
            tax_formset.save()
            items = obj.items.all()
            obj.total_quantity = sum(i.quantity for i in items)
            obj.total_amount = sum(i.amount for i in items)
            obj.total_taxes = sum(t.amount for t in obj.taxes.all())
            obj.grand_total = obj.total_amount + obj.total_taxes
            obj.save()
        messages.success(request, 'PO updated.')
        return redirect('spares:order_detail', pk=pk)
    return render(request, 'spares/po_form.html', {
        'form': form, 'formset': formset, 'tax_formset': tax_formset, 'title': 'Edit PO', 'object': obj
    })


# -- Purchase Invoices --------------------------------------------------------

@login_required
def invoice_list(request):
    invoices = PurchaseInvoice.objects.select_related('supplier').order_by('-date')
    paginator = Paginator(invoices, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/invoice_list.html', {'invoices': page_obj, 'page_obj': page_obj})


@login_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    items = invoice.items.select_related('item', 'warehouse', 'rack', 'bin')
    return render(request, 'spares/invoice_detail.html', {'invoice': invoice, 'items': items})


@login_required
@require_module_action('spares', 'create')
def invoice_create(request):
    form = PurchaseInvoiceForm(request.POST or None)
    formset = PurchaseInvoiceItemFormSet(request.POST or None, prefix='items')
    tax_formset = PurchaseInvoiceTaxFormSet(request.POST or None, prefix='taxes')
    if request.method == 'POST' and form.is_valid() and formset.is_valid() and tax_formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            formset.instance = obj
            formset.save()
            tax_formset.instance = obj
            tax_formset.save()
            all_items = obj.items.all()
            obj.total_quantity = sum(i.quantity for i in all_items)
            obj.total_amount = sum(i.amount for i in all_items)
            obj.total_sgst = sum(i.sgst_amount for i in all_items)
            obj.total_cgst = sum(i.cgst_amount for i in all_items)
            # Interstate suppliers route their whole line GST through
            # igst_amount instead of sgst_amount/cgst_amount (see split_gst()
            # in billing/models.py) -- total_igst must be summed and added to
            # grand_total the same way total_sgst/total_cgst are, otherwise an
            # interstate invoice's grand_total silently drops its entire GST.
            obj.total_igst = sum(i.igst_amount for i in all_items)
            # Bug fix: grand_total used to silently drop the item-level GST
            # (total_sgst/total_cgst, always populated) and only add the
            # separate header "taxes" formset (total_taxes, often empty) --
            # e.g. a Rs.10,000 line at 9%+9% GST produced a grand_total that
            # excluded the Rs.1,800 tax entirely whenever no header tax rows
            # were added. Both are real, additive amounts on this invoice.
            obj.total_taxes = sum(t.amount for t in obj.taxes.all())
            obj.grand_total = obj.total_amount + obj.total_sgst + obj.total_cgst + obj.total_igst + obj.total_taxes
            obj.save()
            # Stock is already credited by PurchaseInvoiceItem.save() itself
            # (it checks is_new and self.invoice.status == 'submitted') --
            # this used to duplicate that same credit a second time here.
        messages.success(request, f'Invoice {obj.invoice_no} created.')
        return redirect('spares:invoice_detail', pk=obj.pk)
    return render(request, 'spares/invoice_form.html', {
        'form': form, 'formset': formset, 'tax_formset': tax_formset, 'title': 'New Purchase Invoice'
    })


@login_required
@require_module_action('spares', 'edit')
def invoice_update(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(PurchaseInvoice, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    # Edit-lock, same convention as order_update: once a PI is no longer
    # Draft, block the full formset rewrite.
    if obj.status != 'draft':
        messages.error(
            request,
            f'{obj.invoice_no} is "{obj.get_status_display()}" and can no longer be edited. '
            'Only a Draft invoice can be changed.'
        )
        return redirect('spares:invoice_detail', pk=pk)
    form = PurchaseInvoiceForm(request.POST or None, instance=obj)
    formset = PurchaseInvoiceItemFormSet(request.POST or None, instance=obj, prefix='items')
    tax_formset = PurchaseInvoiceTaxFormSet(request.POST or None, instance=obj, prefix='taxes')
    if request.method == 'POST' and form.is_valid() and formset.is_valid() and tax_formset.is_valid():
        with transaction.atomic():
            form.save()
            formset.save()
            tax_formset.save()
            all_items = obj.items.all()
            obj.total_quantity = sum(i.quantity for i in all_items)
            obj.total_amount = sum(i.amount for i in all_items)
            obj.total_sgst = sum(i.sgst_amount for i in all_items)
            obj.total_cgst = sum(i.cgst_amount for i in all_items)
            obj.total_igst = sum(i.igst_amount for i in all_items)
            obj.total_taxes = sum(t.amount for t in obj.taxes.all())
            obj.grand_total = obj.total_amount + obj.total_sgst + obj.total_cgst + obj.total_igst + obj.total_taxes
            obj.save()
        messages.success(request, 'Invoice updated.')
        return redirect('spares:invoice_detail', pk=pk)
    return render(request, 'spares/invoice_form.html', {
        'form': form, 'formset': formset, 'tax_formset': tax_formset, 'title': 'Edit Invoice', 'object': obj
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def invoice_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(PurchaseInvoice, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Purchase Invoice', 'update', pk)
        messages.success(request, f'{obj.invoice_no} submitted.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:invoice_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def invoice_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(PurchaseInvoice, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Purchase Invoice', 'update', pk)
        messages.success(request, f'{obj.invoice_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:invoice_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'create')
def invoice_amend(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(PurchaseInvoice, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        new = obj.amend()
        new.created_by = request.user
        new.save(update_fields=['created_by'])
        log_action(request, 'Purchase Invoice', 'create', new.pk)
        messages.success(request, f'{new.invoice_no} created as an amendment of {obj.invoice_no}.')
        return redirect('spares:invoice_update', pk=new.pk)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect('spares:invoice_detail', pk=pk)


# -- Counter Sales ------------------------------------------------------------

@login_required
def counter_sale_list(request):
    sales = CounterSale.objects.select_related('godown').order_by('-date')
    paginator = Paginator(sales, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/counter_sale_list.html', {'sales': page_obj, 'page_obj': page_obj})


@login_required
def counter_sale_detail(request, pk):
    sale = get_object_or_404(CounterSale, pk=pk)
    items = sale.items.select_related('item', 'rack', 'bin')
    return render(request, 'spares/counter_sale_detail.html', {'sale': sale, 'items': items})


@login_required
@require_module_action('spares', 'create')
def counter_sale_create(request):
    form = CounterSaleForm(request.POST or None)
    formset = CounterSaleItemFormSet(request.POST or None, prefix='items')
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        requested_status = form.cleaned_data.get('status') or 'draft'
        try:
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.created_by = request.user
                # Force Draft on this first save regardless of what the form
                # requested -- items don't exist yet at this point, and the
                # StockLedger-posting signal on CounterSale fires on every
                # save(). Posting "submitted" here would run the movement
                # loop against zero items and permanently mark stock_posted,
                # skipping the real decrement once items are actually added.
                obj.status = 'draft'
                obj.save()
                formset.instance = obj
                formset.save()
                all_items = obj.items.all()
                obj.total_qty = sum(i.quantity for i in all_items)
                obj.total_amount = sum(i.total for i in all_items) - obj.discount_amount
                obj.save()
                if requested_status == 'submitted':
                    # Stock-safety guard lives in obj.submit(): raises
                    # ValueError (caught below) if any line would drive
                    # StockLedger negative, before the decrement is posted.
                    obj.submit(request.user)
                elif requested_status == 'cancelled':
                    obj.status = 'cancelled'
                    obj.save()
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, 'spares/counter_sale_form.html', {
                'form': form, 'formset': formset, 'title': 'New Counter Sale'
            })
        messages.success(request, f'Counter sale {obj.sale_no} created.')
        return redirect('spares:counter_sale_detail', pk=obj.pk)
    return render(request, 'spares/counter_sale_form.html', {
        'form': form, 'formset': formset, 'title': 'New Counter Sale'
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def counter_sale_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(CounterSale, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Counter Sale', 'update', pk)
        messages.success(request, f'{obj.sale_no} submitted.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:counter_sale_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def counter_sale_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(CounterSale, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Counter Sale', 'update', pk)
        messages.success(request, f'{obj.sale_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:counter_sale_detail', pk=pk)


# -- Counter Returns ----------------------------------------------------------

@login_required
def counter_return_list(request):
    returns = CounterSaleReturn.objects.select_related('original_sale').order_by('-return_date')
    paginator = Paginator(returns, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/counter_return_list.html', {'returns': page_obj, 'page_obj': page_obj})


@login_required
def counter_return_detail(request, pk):
    ret = get_object_or_404(CounterSaleReturn, pk=pk)
    items = ret.items.select_related('item')
    return render(request, 'spares/counter_return_detail.html', {'ret': ret, 'items': items})


@login_required
@require_module_action('spares', 'create')
def counter_return_create(request):
    form = CounterSaleReturnForm(request.POST or None)
    formset = CounterSaleReturnItemFormSet(request.POST or None, prefix='items')
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            # status isn't exposed on this form (system-controlled) -- stays
            # at the model default 'draft' until items exist, then this view
            # auto-submits it immediately below (a return has always behaved
            # as a single-step, already-final document in this app's UI; no
            # stock-safety check needed here since a return only adds stock).
            obj.save()
            formset.instance = obj
            formset.save()
            all_items = obj.items.all()
            obj.total_amount = sum(i.amount for i in all_items)
            obj.save()
            obj.submit(request.user)
        messages.success(request, f'Return {obj.return_no} created.')
        return redirect('spares:counter_return_detail', pk=obj.pk)
    return render(request, 'spares/counter_return_form.html', {
        'form': form, 'formset': formset, 'title': 'New Counter Return'
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def counter_return_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(CounterSaleReturn, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Counter Sale Return', 'update', pk)
        messages.success(request, f'{obj.return_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:counter_return_detail', pk=pk)


# -- Issue Alterations --------------------------------------------------------

@login_required
def issue_alteration_list(request):
    alterations = SparesIssueAlteration.objects.select_related('godown').order_by('-date')
    paginator = Paginator(alterations, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/issue_alteration_list.html', {'alterations': page_obj, 'page_obj': page_obj})


@login_required
def issue_alteration_detail(request, pk):
    alteration = get_object_or_404(SparesIssueAlteration, pk=pk)
    items = alteration.items.select_related('item', 'rack', 'bin')
    deleted_items = alteration.deleted_items.select_related('item', 'rack', 'bin')
    return render(request, 'spares/issue_alteration_detail.html', {
        'alteration': alteration, 'items': items, 'deleted_items': deleted_items,
        'returns': alteration.returns.all(),
    })


@login_required
@require_module_action('spares', 'create')
def issue_alteration_create(request):
    initial = {}
    if request.GET.get('jc'):
        initial['job_card'] = request.GET['jc']
    if request.GET.get('uv_jc'):
        initial['used_vehicle_job_card'] = request.GET['uv_jc']
    form = SparesIssueAlterationForm(request.POST or None, initial=initial)
    formset = SparesIssueAlterationItemFormSet(request.POST or None, prefix='items')
    deleted_formset = SparesIssueAlterationDeletedItemFormSet(request.POST or None, prefix='deleted')
    if request.method == 'POST' and form.is_valid() and formset.is_valid() and deleted_formset.is_valid():
        try:
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.created_by = request.user
                # status isn't exposed on this form (system-controlled) --
                # stays at the model default 'draft' until items exist, then
                # this view auto-submits it below with a stock-safety check
                # (spares issued to a job card leave inventory, so unlike a
                # return, insufficient stock must block the submission).
                obj.save()
                formset.instance = obj
                formset.save()
                deleted_formset.instance = obj
                deleted_formset.save()
                all_items = obj.items.all()
                obj.spares_total = sum(i.total for i in all_items)
                obj.total = obj.spares_total + obj.labour_total + obj.outwork_total
                obj.updated_total = obj.total - obj.discount
                obj.save()
                obj.submit(request.user)
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, 'spares/issue_alteration_form.html', {
                'form': form, 'formset': formset, 'deleted_formset': deleted_formset, 'title': 'New Issue Alteration'
            })
        messages.success(request, f'Issue alteration SIA-{obj.pk:05d} created.')
        return redirect('spares:issue_alteration_detail', pk=obj.pk)
    return render(request, 'spares/issue_alteration_form.html', {
        'form': form, 'formset': formset, 'deleted_formset': deleted_formset, 'title': 'New Issue Alteration'
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def issue_alteration_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(SparesIssueAlteration, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Spares Issue Alteration', 'update', pk)
        messages.success(request, f'SIA-{obj.pk:05d} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:issue_alteration_detail', pk=pk)


# -- AJAX ---------------------------------------------------------------------

@login_required
def ajax_item_details(request):
    item_id = request.GET.get('item_id')
    if not item_id:
        return JsonResponse({'error': 'No item_id'}, status=400)
    try:
        item = SparesItem.objects.get(pk=item_id)
        return JsonResponse({
            'rate': float(item.standard_selling_rate),
            'mrp': float(item.mrp),
            'sgst': float(item.sgst),
            'cgst': float(item.cgst),
            'uom': item.uom,
            'hsn': item.hsn_sac,
        })
    except SparesItem.DoesNotExist:
        return JsonResponse({'error': 'Item not found'}, status=404)


@login_required
def ajax_supplier_details(request):
    supplier_id = request.GET.get('supplier_id')
    if not supplier_id:
        return JsonResponse({'error': 'No supplier_id'}, status=400)
    try:
        supplier = Supplier.objects.get(pk=supplier_id)
        return JsonResponse({
            'gstin': supplier.gstin,
            'gst_category': supplier.gst_category,
            'place_of_supply': supplier.place_of_supply,
            'address': ', '.join(filter(None, [
                supplier.address_line1, supplier.address_line2,
                supplier.city, supplier.state, supplier.pincode
            ])),
        })
    except Supplier.DoesNotExist:
        return JsonResponse({'error': 'Supplier not found'}, status=404)


@login_required
def ajax_rack_bins(request):
    rack_id = request.GET.get('rack_id')
    if not rack_id:
        return JsonResponse({'bins': []})
    bins = list(Bin.objects.filter(rack_id=rack_id).values('id', 'name'))
    return JsonResponse({'bins': bins})


# ---------------------------------------------------------------------------
# GAP 13: Bulk Insert (CSV)
# ---------------------------------------------------------------------------

@login_required
def bulk_insert(request):
    import csv
    from io import TextIOWrapper, StringIO
    from decimal import Decimal, InvalidOperation

    MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5MB

    results = None
    if request.method == 'POST' and request.FILES.get('file'):
        f = request.FILES['file']
        if f.size > MAX_UPLOAD_BYTES:
            messages.error(request, 'File is too large — the limit is 5MB.')
            return render(request, 'spares/bulk_insert.html', {'results': None})
        name = (f.name or '').lower()
        rows = []
        errors = []
        imported = []
        try:
            if name.endswith('.csv') or name.endswith('.txt'):
                text = TextIOWrapper(f.file, encoding='utf-8-sig', errors='ignore')
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith('.xlsx') or name.endswith('.xls'):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(f, read_only=True, data_only=True)
                    ws = wb.active
                    headers = None
                    for r in ws.iter_rows(values_only=True):
                        if headers is None:
                            headers = [str(c).strip() if c is not None else '' for c in r]
                            continue
                        if all(c is None for c in r):
                            continue
                        rows.append({headers[i]: r[i] for i in range(len(headers))})
                except ImportError:
                    errors.append({'row': 0, 'error': 'openpyxl not installed; please upload CSV.'})
            else:
                errors.append({'row': 0, 'error': 'Unsupported file type. Use .csv or .xlsx'})
        except Exception as exc:
            errors.append({'row': 0, 'error': f'Could not read file: {exc}'})

        def D(v, default='0'):
            if v in (None, ''):
                return Decimal(default)
            try:
                return Decimal(str(v).strip())
            except (InvalidOperation, ValueError):
                return Decimal(default)

        def _defuse(value):
            # Prevent CSV/Excel formula injection if this data is ever
            # re-exported and opened by someone else.
            if value and value[0] in ('=', '+', '-', '@'):
                return "'" + value
            return value

        for idx, row in enumerate(rows, start=2):
            try:
                row = {(k or '').strip().lower(): v for k, v in row.items()}
                name_v = _defuse((row.get('item_name') or '').strip())
                if not name_v:
                    errors.append({'row': idx, 'error': 'item_name is required'})
                    continue
                part_no = _defuse((row.get('part_number') or '').strip())
                hsn = (row.get('hsn_sac') or '').strip()
                category = _defuse((row.get('category') or '').strip())

                item = SparesItem(
                    item_name=name_v,
                    part_number=part_no,
                    hsn_sac=hsn,
                    mrp=D(row.get('mrp')),
                    standard_selling_rate=D(row.get('selling_rate')),
                    sgst=D(row.get('sgst'), '9'),
                    cgst=D(row.get('cgst'), '9'),
                    reorder_level=D(row.get('reorder_level')),
                )
                item.save()
                if category:
                    try:
                        from masters.models import SparesCategory
                        cat, _ = SparesCategory.objects.get_or_create(name=category)
                        item.category = cat
                        item.save(update_fields=['category'])
                    except Exception:
                        pass
                imported.append({'row': idx, 'item_code': item.item_code, 'name': item.item_name})
            except Exception as exc:
                errors.append({'row': idx, 'error': str(exc)})

        results = {
            'total': len(rows),
            'imported': imported,
            'errors': errors,
            'imported_count': len(imported),
            'error_count': len(errors),
        }
        if imported:
            messages.success(request, f"{len(imported)} items imported successfully.")
        if errors:
            messages.warning(request, f"{len(errors)} rows had errors.")

    return render(request, 'spares/bulk_insert.html', {'results': results})


# ---------------------------------------------------------------------------
# GAP 21 — PO Used Qty Report
# ---------------------------------------------------------------------------

@login_required
def po_used_qty_report(request):
    from .models import (CounterSaleItem, PurchaseOrder, PurchaseOrderItem,
                         SparesIssueAlterationItem)

    items_qs = list(
        PurchaseOrderItem.objects.select_related('order__supplier', 'item').order_by('-order__date')
    )
    item_ids = [poi.item_id for poi in items_qs]

    # Two batched aggregate queries instead of two per row (N+1).
    used_by_item = dict(
        SparesIssueAlterationItem.objects.filter(item_id__in=item_ids)
        .values('item_id').annotate(t=Sum('quantity')).values_list('item_id', 't')
    )
    sold_by_item = dict(
        CounterSaleItem.objects.filter(item_id__in=item_ids)
        .values('item_id').annotate(t=Sum('quantity')).values_list('item_id', 't')
    )

    rows = []
    for poi in items_qs:
        item_id = poi.item_id
        ordered = poi.quantity or 0
        used_in_service = used_by_item.get(item_id) or 0
        sold_counter = sold_by_item.get(item_id) or 0
        consumed = (used_in_service or 0) + (sold_counter or 0)
        remaining = (ordered or 0) - consumed
        rows.append({
            'po_no': poi.order.po_no,
            'supplier': poi.order.supplier.supplier_name if poi.order.supplier else '—',
            'item': poi.item,
            'ordered': ordered,
            'used_service': used_in_service,
            'sold_counter': sold_counter,
            'remaining': remaining,
        })
    return render(request, 'spares/po_used_qty_report.html', {'rows': rows})


# ============================================================
# FEATURE 8 — Parts Consumption Report
# ============================================================

@login_required
def parts_consumption_report(request):
    from django.db.models import Sum
    from datetime import date
    import calendar as _cal
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year  = int(request.GET.get('year',  today.year))

    service_consumption = SparesIssueAlterationItem.objects.filter(
        alteration__date__month=month,
        alteration__date__year=year
    ).values(
        'item__item_name', 'item__item_code', 'item__category__name'
    ).annotate(
        total_qty=Sum('quantity'), total_value=Sum('total')
    ).order_by('-total_qty')

    counter_consumption = CounterSaleItem.objects.filter(
        sale__date__month=month,
        sale__date__year=year,
        sale__status='submitted'
    ).values(
        'item__item_name', 'item__item_code', 'item__category__name'
    ).annotate(
        total_qty=Sum('quantity'), total_value=Sum('total')
    ).order_by('-total_qty')

    context = {
        'month': month, 'year': year,
        'month_name': _cal.month_name[month],
        'service_consumption': service_consumption,
        'counter_consumption': counter_consumption,
    }
    return render(request, 'spares/parts_consumption_report.html', context)


# ---------------------------------------------------------------------------
# Delete views
# ---------------------------------------------------------------------------
from django.views.decorators.http import require_POST


@login_required
@require_POST
@require_module_action('spares', 'delete')
def item_delete(request, pk):
    from django.db.models import ProtectedError
    item = get_object_or_404(SparesItem, pk=pk)
    try:
        item.delete()
        log_action(request, 'SparesItem', 'delete', pk)
        messages.success(request, f'Item deleted.')
    except ProtectedError:
        messages.error(
            request,
            f'Cannot delete "{item.item_name}": it is referenced by purchase orders, '
            'invoices, or sales records. Deactivate it instead.'
        )
        return redirect('spares:item_detail', pk=pk)
    return redirect('spares:item_list')


@login_required
@require_POST
@require_module_action('spares', 'delete')
def purchase_order_delete(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(PurchaseOrder, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    if obj.status in ('received', 'submitted'):
        messages.error(
            request,
            f'Cannot delete {obj.po_no}: status is "{obj.get_status_display()}". '
            'Only Draft or Cancelled orders can be deleted.'
        )
        return redirect('spares:order_detail', pk=pk)
    obj.delete()
    log_action(request, 'PurchaseOrder', 'delete', pk)
    messages.success(request, f'Purchase order {obj.po_no} deleted.')
    return redirect('spares:order_list')


# ---------------------------------------------------------------------------
# Phase 7a — Stock Transfer
# ---------------------------------------------------------------------------

@login_required
def stock_transfer_list(request):
    transfers = StockTransfer.objects.select_related('warehouse', 'branch').order_by('-date_and_time')
    paginator = Paginator(transfers, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/stock_transfer_list.html', {'transfers': page_obj, 'page_obj': page_obj})


@login_required
def stock_transfer_detail(request, pk):
    obj = get_object_or_404(StockTransfer.objects.select_related('warehouse', 'branch'), pk=pk)
    items = obj.items.select_related('item', 'from_rack', 'from_bin', 'to_warehouse', 'to_rack', 'to_bin')
    return render(request, 'spares/stock_transfer_detail.html', {'obj': obj, 'items': items})


@login_required
@require_module_action('spares', 'create')
def stock_transfer_create(request):
    form = StockTransferForm(request.POST or None)
    formset = StockTransferItemFormSet(request.POST or None, prefix='items')
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            formset.instance = obj
            formset.save()
        log_action(request, 'Stock Transfer', 'create', obj.pk)
        messages.success(request, f'{obj.transfer_no} created.')
        return redirect('spares:stock_transfer_detail', pk=obj.pk)
    return render(request, 'spares/stock_transfer_form.html', {
        'form': form, 'formset': formset, 'title': 'New Stock Transfer',
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def stock_transfer_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(StockTransfer, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Stock Transfer', 'update', pk)
        messages.success(request, f'{obj.transfer_no} submitted.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:stock_transfer_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def stock_transfer_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(StockTransfer, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Stock Transfer', 'update', pk)
        messages.success(request, f'{obj.transfer_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:stock_transfer_detail', pk=pk)


# ---------------------------------------------------------------------------
# Phase 7a — Stock Count Update (Spares Stock Reconciliation)
# ---------------------------------------------------------------------------

@login_required
def stock_count_list(request):
    counts = StockCountUpdate.objects.select_related('warehouse', 'branch').order_by('-date_and_time')
    paginator = Paginator(counts, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/stock_count_list.html', {'counts': page_obj, 'page_obj': page_obj})


@login_required
def stock_count_detail(request, pk):
    obj = get_object_or_404(StockCountUpdate.objects.select_related('warehouse', 'branch'), pk=pk)
    items = obj.items.select_related('item', 'rack', 'bin')
    return render(request, 'spares/stock_count_detail.html', {'obj': obj, 'items': items})


@login_required
@require_module_action('spares', 'create')
def stock_count_create(request):
    form = StockCountUpdateForm(request.POST or None)
    formset = StockCountItemFormSet(request.POST or None, prefix='items')
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            formset.instance = obj
            formset.save()
        log_action(request, 'Stock Count Update', 'create', obj.pk)
        messages.success(request, f'{obj.count_no} created.')
        return redirect('spares:stock_count_detail', pk=obj.pk)
    return render(request, 'spares/stock_count_form.html', {
        'form': form, 'formset': formset, 'title': 'New Stock Count Update',
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def stock_count_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(StockCountUpdate, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Stock Count Update', 'update', pk)
        messages.success(request, f'{obj.count_no} submitted.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:stock_count_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def stock_count_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(StockCountUpdate, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Stock Count Update', 'update', pk)
        messages.success(request, f'{obj.count_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:stock_count_detail', pk=pk)


# ---------------------------------------------------------------------------
# Phase 7a — Request Supplier Quote
# ---------------------------------------------------------------------------

@login_required
def request_supplier_quote_list(request):
    objs = RequestSupplierQuote.objects.prefetch_related('suppliers').order_by('-date')
    paginator = Paginator(objs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/request_supplier_quote_list.html', {'objs': page_obj, 'page_obj': page_obj})


@login_required
def request_supplier_quote_detail(request, pk):
    obj = get_object_or_404(RequestSupplierQuote.objects.prefetch_related('suppliers'), pk=pk)
    items = obj.items.select_related('spare')
    return render(request, 'spares/request_supplier_quote_detail.html', {
        'obj': obj, 'items': items, 'supplier_quotes': obj.supplier_quotes.all(),
    })


@login_required
@require_module_action('spares', 'create')
def request_supplier_quote_create(request):
    form = RequestSupplierQuoteForm(request.POST or None)
    formset = RequestSupplierQuoteItemFormSet(request.POST or None, prefix='items')
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            form.save_m2m()
            formset.instance = obj
            formset.save()
        log_action(request, 'Request Supplier Quote', 'create', obj.pk)
        messages.success(request, f'{obj.rsq_no} created.')
        return redirect('spares:request_supplier_quote_detail', pk=obj.pk)
    return render(request, 'spares/request_supplier_quote_form.html', {
        'form': form, 'formset': formset, 'title': 'New Request Supplier Quote',
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def request_supplier_quote_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(RequestSupplierQuote, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Request Supplier Quote', 'update', pk)
        messages.success(request, f'{obj.rsq_no} submitted.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:request_supplier_quote_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def request_supplier_quote_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(RequestSupplierQuote, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Request Supplier Quote', 'update', pk)
        messages.success(request, f'{obj.rsq_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:request_supplier_quote_detail', pk=pk)


# ---------------------------------------------------------------------------
# Phase 7b — Spares Purchase Estimation Master
# ---------------------------------------------------------------------------

@login_required
def estimation_list(request):
    objs = SparesPurchaseEstimationMaster.objects.select_related('insurance_name').order_by('-date')
    paginator = Paginator(objs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/estimation_list.html', {'objs': page_obj, 'page_obj': page_obj})


@login_required
def estimation_detail(request, pk):
    obj = get_object_or_404(SparesPurchaseEstimationMaster.objects.select_related('insurance_name'), pk=pk)
    items = obj.items.select_related('item')
    labor = obj.labor_details.select_related('labor_name')
    return render(request, 'spares/estimation_detail.html', {
        'obj': obj, 'items': items, 'labor': labor,
        'purchase_orders': obj.purchase_orders.all(),
    })


@login_required
@require_module_action('spares', 'create')
def estimation_create(request):
    form = SparesPurchaseEstimationMasterForm(request.POST or None)
    item_formset = SparesPurchaseEstimationItemFormSet(request.POST or None, prefix='items')
    labor_formset = SparesPurchaseEstimationLaborFormSet(request.POST or None, prefix='labor')
    if request.method == 'POST' and form.is_valid() and item_formset.is_valid() and labor_formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            item_formset.instance = obj
            item_formset.save()
            labor_formset.instance = obj
            labor_formset.save()
            items = obj.items.all()
            labor_lines = obj.labor_details.all()
            obj.estimation_total_amount = sum(i.total for i in items)
            obj.balance_delivery_qty = sum(i.delivery_balance_qty for i in items)
            obj.labor_total_amount = sum(l.total for l in labor_lines)
            obj.total_amount = obj.estimation_total_amount + obj.labor_total_amount
            obj.save()
        log_action(request, 'Spares Purchase Estimation Master', 'create', obj.pk)
        messages.success(request, f'{obj.estimation_no} created.')
        return redirect('spares:estimation_detail', pk=obj.pk)
    return render(request, 'spares/estimation_form.html', {
        'form': form, 'item_formset': item_formset, 'labor_formset': labor_formset,
        'title': 'New Spares Purchase Estimation',
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def estimation_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(SparesPurchaseEstimationMaster, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Spares Purchase Estimation Master', 'update', pk)
        messages.success(request, f'{obj.estimation_no} submitted.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:estimation_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def estimation_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(SparesPurchaseEstimationMaster, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Spares Purchase Estimation Master', 'update', pk)
        messages.success(request, f'{obj.estimation_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:estimation_detail', pk=pk)


# ---------------------------------------------------------------------------
# Phase 7c — Service Spares Issue Return
# ---------------------------------------------------------------------------

@login_required
def service_spares_issue_return_list(request):
    objs = ServiceSparesIssueReturn.objects.select_related('spares_issue', 'job_card').order_by('-created_at')
    paginator = Paginator(objs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/service_spares_issue_return_list.html', {'objs': page_obj, 'page_obj': page_obj})


@login_required
def service_spares_issue_return_detail(request, pk):
    obj = get_object_or_404(
        ServiceSparesIssueReturn.objects.select_related('spares_issue', 'job_card', 'godown'), pk=pk
    )
    items = obj.items.select_related('item', 'rack', 'bin')
    return render(request, 'spares/service_spares_issue_return_detail.html', {'obj': obj, 'items': items})


@login_required
@require_module_action('spares', 'create')
def service_spares_issue_return_create(request):
    initial = {}
    if request.GET.get('spares_issue'):
        initial['spares_issue'] = request.GET['spares_issue']
    form = ServiceSparesIssueReturnForm(request.POST or None, initial=initial)
    formset = ServiceSparesIssueReturnItemFormSet(request.POST or None, prefix='items')
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            formset.instance = obj
            formset.save()
            items = obj.items.all()
            obj.total_qty = sum(i.return_qty for i in items)
            obj.total_amount = sum(i.total for i in items)
            obj.save()
        log_action(request, 'Service Spares Issue Return', 'create', obj.pk)
        messages.success(request, f'{obj.return_no} created.')
        return redirect('spares:service_spares_issue_return_detail', pk=obj.pk)
    return render(request, 'spares/service_spares_issue_return_form.html', {
        'form': form, 'formset': formset, 'title': 'New Service Spares Issue Return',
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def service_spares_issue_return_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(ServiceSparesIssueReturn, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Service Spares Issue Return', 'update', pk)
        messages.success(request, f'{obj.return_no} submitted.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:service_spares_issue_return_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def service_spares_issue_return_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(ServiceSparesIssueReturn, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Service Spares Issue Return', 'update', pk)
        messages.success(request, f'{obj.return_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:service_spares_issue_return_detail', pk=pk)


# ---------------------------------------------------------------------------
# Phase 7d — Vehicle Spares Master
# ---------------------------------------------------------------------------

@login_required
def vehicle_spares_master_list(request):
    objs = VehicleSparesMaster.objects.select_related('spare').order_by('spare__item_name')
    paginator = Paginator(objs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/vehicle_spares_master_list.html', {'objs': page_obj, 'page_obj': page_obj})


@login_required
@require_module_action('spares', 'create')
def vehicle_spares_master_create(request):
    form = VehicleSparesMasterForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save()
        log_action(request, 'Vehicle Spares Master', 'create', obj.pk)
        messages.success(request, f'Vehicle Spares Master for {obj.spare.item_code} saved.')
        return redirect('spares:vehicle_spares_master_list')
    return render(request, 'spares/vehicle_spares_master_form.html', {
        'form': form, 'title': 'New Vehicle Spares Master',
    })


# ---------------------------------------------------------------------------
# Phase 7d — Spares MRP Price Revision
# ---------------------------------------------------------------------------

@login_required
def mrp_revision_list(request):
    objs = SparesMRPPriceRevision.objects.order_by('-date')
    paginator = Paginator(objs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/mrp_revision_list.html', {'objs': page_obj, 'page_obj': page_obj})


@login_required
def mrp_revision_detail(request, pk):
    obj = get_object_or_404(SparesMRPPriceRevision, pk=pk)
    items = obj.items.select_related('item')
    return render(request, 'spares/mrp_revision_detail.html', {'obj': obj, 'items': items})


@login_required
@require_module_action('spares', 'create')
def mrp_revision_create(request):
    form = SparesMRPPriceRevisionForm(request.POST or None)
    formset = SparesMRPPriceRevisionItemFormSet(request.POST or None, prefix='items')
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            formset.instance = obj
            formset.save()
        log_action(request, 'Spares MRP Price Revision', 'create', obj.pk)
        messages.success(request, f'{obj.revision_no} created.')
        return redirect('spares:mrp_revision_detail', pk=obj.pk)
    return render(request, 'spares/mrp_revision_form.html', {
        'form': form, 'formset': formset, 'title': 'New Spares MRP Price Revision',
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def mrp_revision_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(SparesMRPPriceRevision, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Spares MRP Price Revision', 'update', pk)
        messages.success(request, f'{obj.revision_no} submitted — item prices updated.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:mrp_revision_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def mrp_revision_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(SparesMRPPriceRevision, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Spares MRP Price Revision', 'update', pk)
        messages.success(request, f'{obj.revision_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:mrp_revision_detail', pk=pk)


# ---------------------------------------------------------------------------
# Phase 7d — Spares settings singles (Profit Percentage / Purchase Qty Days)
# ---------------------------------------------------------------------------

@login_required
def spares_settings(request):
    profit_instance = SparesProfitPercentageSettings.get_instance()
    qty_days_instance = SparesPurchaseQtyDaysSettings.get_instance()
    is_profit_save = request.method == 'POST' and 'save_profit' in request.POST
    is_qty_days_save = request.method == 'POST' and 'save_qty_days' in request.POST
    profit_form = SparesProfitPercentageSettingsForm(
        request.POST if is_profit_save else None, instance=profit_instance, prefix='profit',
    )
    qty_days_form = SparesPurchaseQtyDaysSettingsForm(
        request.POST if is_qty_days_save else None, instance=qty_days_instance, prefix='qty_days',
    )
    if is_profit_save and profit_form.is_valid():
        profit_form.save()
        messages.success(request, 'Spares Profit Percentage settings saved.')
        return redirect('spares:spares_settings')
    if is_qty_days_save and qty_days_form.is_valid():
        qty_days_form.save()
        messages.success(request, 'Spares Purchase Order Qty Days settings saved.')
        return redirect('spares:spares_settings')
    return render(request, 'spares/spares_settings.html', {
        'profit_form': profit_form, 'qty_days_form': qty_days_form,
    })


# ---------------------------------------------------------------------------
# Phase 7d — Service Spares Warranty
# ---------------------------------------------------------------------------

@login_required
def service_spares_warranty_list(request):
    objs = ServiceSparesWarranty.objects.select_related('supplier').order_by('-claim_date')
    paginator = Paginator(objs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'spares/service_spares_warranty_list.html', {'objs': page_obj, 'page_obj': page_obj})


@login_required
def service_spares_warranty_detail(request, pk):
    obj = get_object_or_404(ServiceSparesWarranty.objects.select_related('supplier', 'branch'), pk=pk)
    items = obj.items.select_related('item')
    return render(request, 'spares/service_spares_warranty_detail.html', {'obj': obj, 'items': items})


@login_required
@require_module_action('spares', 'create')
def service_spares_warranty_create(request):
    form = ServiceSparesWarrantyForm(request.POST or None)
    formset = ServiceSparesWarrantyItemFormSet(request.POST or None, prefix='items')
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            formset.instance = obj
            formset.save()
        log_action(request, 'Service Spares Warranty', 'create', obj.pk)
        messages.success(request, f'{obj.warranty_no} created.')
        return redirect('spares:service_spares_warranty_detail', pk=obj.pk)
    return render(request, 'spares/service_spares_warranty_form.html', {
        'form': form, 'formset': formset, 'title': 'New Service Spares Warranty',
    })


@login_required
@require_POST
@require_module_action('spares', 'edit')
def service_spares_warranty_submit(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(ServiceSparesWarranty, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.submit(request.user)
        log_action(request, 'Service Spares Warranty', 'update', pk)
        messages.success(request, f'{obj.warranty_no} submitted.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:service_spares_warranty_detail', pk=pk)


@login_required
@require_POST
@require_module_action('spares', 'edit')
def service_spares_warranty_cancel(request, pk):
    from accounts.permissions import user_owns
    from django.http import HttpResponseForbidden
    obj = get_object_or_404(ServiceSparesWarranty, pk=pk)
    if not user_owns(request.user, obj):
        return HttpResponseForbidden('<h1>403 — Access Denied</h1>')
    try:
        obj.cancel(request.user)
        log_action(request, 'Service Spares Warranty', 'update', pk)
        messages.success(request, f'{obj.warranty_no} cancelled.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('spares:service_spares_warranty_detail', pk=pk)
